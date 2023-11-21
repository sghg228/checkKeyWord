
import os

from chardet.universaldetector import UniversalDetector
import getpass
import docx2txt
from openpyxl import load_workbook
import time
import socket
from striprtf.striprtf import rtf_to_text

start_time = time.time()

def check_keywords_in_text(text, keywords):
    found_keywords = [keyword for keyword in keywords if keyword.lower() in text.lower()]
    return found_keywords


def check_keywords_in_docx(file_path, keywords):
    content = ''
    try:
        content = docx2txt.process(file_path)
    except Exception as e:
        k = 1
        print(f'Ошибка: {e} {file_path}')
    return check_keywords_in_text(content, keywords)


def detection_encoding(file_path):
    try:
        detector = UniversalDetector()
        with open(file_path, 'r') as fh:
            for line in fh:
                detector.feed(line)
                if detector.done:
                    break
            detector.close()

    except Exception as e4:
        print(f'Ошибка: {e4} {file_path} {detector.result["encoding"]}')

    if(detector.result["encoding"] == None ):
        return "utf-8"
    print(detector.result["encoding"])
    return detector.result["encoding"]


def check_keywords_in_rtf(file_path, keywords):
    text = ''
    found_keywords = None
    try:
        with open(file_path) as infile:
            content = infile.read()
            text = rtf_to_text(content)
    except Exception as e:
        try:
            with open(file_path, encoding='cp1251') as infile:
                content = infile.read()
                text = rtf_to_text(content)
        except Exception as e2:
            try:
                with open(file_path, encoding='latin-1') as infile:
                    content = infile.read()
                    text = rtf_to_text(content)
            except Exception as e3:
                found_keywords = check_keywords_in_text(open(file_path, 'r', encoding=str(detection_encoding(file_path)), errors="ignore").read(), keywords)
                print(f'Ошибка: {e3} {file_path}')

    return found_keywords


def check_keywords_in_xlsx(file_path: str, keywords: list) -> str:
    content = []
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    content.append(str(cell.value))
    except Exception as e:
        k = 1
      #  print(str('Ошибка:\n' + str(traceback.format_exc()) + ' ' + str(file_path) + ' КОД ' + str(e)))
    return check_keywords_in_text('\n'.join(content), keywords)


def check_keyword_in_txt(file_path, keywords):
    found_keywords = None
    try:
        found_keywords = check_keywords_in_text(open(file_path, 'r').read(), keywords)
    except Exception as e:
        #print(f'Ошибка: {e} {file_path}')
        try:
            found_keywords = check_keywords_in_text(open(file_path, 'r', encoding="UTF-16").read(), keywords)
        except Exception as e2:
            try:
                found_keywords = check_keywords_in_text(open(file_path, 'r', encoding="utf-8").read(), keywords)
            except Exception as e3:

                try:
                    detector = UniversalDetector()
                    with open(file_path, 'r', errors="ignore") as fh:
                        for line in fh:
                            detector.feed(line)
                            if detector.done:
                                break
                        detector.close()
                    found_keywords = check_keywords_in_text(open(file_path, 'r', errors="ignore", encoding=str(detector.result["encoding"])).read())
                except Exception as e4:
                    print(f'Ошибка: {e3} {file_path} {detector.result["encoding"]}')
                    k = 1

    return found_keywords


def extension_processing(file_path, keywords):
    found_keywords = None
    if file_path.endswith(('.txt', '.docx', '.doc', '.rtf', 'xls', 'xlsx')):
        if file_path.endswith('.txt'):
            found_keywords = check_keyword_in_txt(file_path, keywords)
        elif file_path.endswith(('.docx', '.doc')):
            found_keywords = check_keywords_in_docx(file_path, keywords)
        elif file_path.endswith('.rtf'):
            found_keywords = check_keywords_in_rtf(file_path, keywords)
        elif file_path.endswith('.xlsx'):
            found_keywords = check_keywords_in_xlsx(file_path, keywords)
        return found_keywords
    return found_keywords


def search_files_in_folder(folder_path, keywords):
    name = os.environ['USERNAME']
    log = [name]
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:

            file_path = os.path.join(root, file_name)
            if file_path.endswith(('.zip', '.rar', '.7z')):
                continue
            else:
                found_keywords = extension_processing(file_path, keywords)
            if found_keywords:
                add_log(name, file_path, found_keywords, log)
    return log


def add_log(name, file_path, found_keywords, log):
   log.append(f'{file_path} {found_keywords}')

if __name__ == "__main__":
    # Путь к папке, которую нужно проверить
    folder_path = 'D:\\'
    # Список ключевых слов для поиска
    keywords = ['TEST A', 'TEST B', 'test c', "sghg"]

    log = search_files_in_folder(folder_path, keywords)

    print([str(x) + ' ' for x in log])

end_time = time.time()
# Рассчитайте разницу, чтобы узнать время выполнения
execution_time = end_time - start_time
print(str(execution_time))

with open("D:\\projects\\4year\\Pract\\checkKeyWord\\dist\\out.txt", "w") as file:
    for one_log in log:
        file.write(one_log+"\n")

######## сокеты
